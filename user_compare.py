import openpyxl as op


def get_data(file_name, min_row, max_row):
    """
    функция принимает имя файла, номера строк которые нужно обработать,
    и возвращает список
    :param file_name: ередается имя файла (с расширением) для прочтения
    :param min_row: строка с которой начинается чтение
    :param max_row: строка на которой чтение файла заканчивается
    :return:
    """
    data_list = []
    sheets = op.load_workbook(filename=file_name).active
    for rows in sheets.iter_rows(min_row=min_row, max_row=max_row, min_col=0, max_col=7):
        x = []
        for cell in rows:
            x.append(cell.value)
        data_list.append(x)
    return data_list


def joiner(array):
    """
    принимает списпок, объединяет в строку элементы списка, удаляет
    объединенные в строку элементы, вставляет новое значение за место удаленных
    :param array: список
    :return: список с объединенными элементами списка
    """
    n = 0
    for i in array:
        if i[2] is None: # если колонка пустая - её не обрабатывать.
            x = i[1] + ' ' + i[3]
        else:
            x = i[3] + ' ' + i[1] + ' ' + i[2]
        del array[n][1:4]
        array[n].insert(1, x)
        n += 1
    return array


active_users = get_data('active2.xlsx', 1, 421) # 7717-SD,  3815-active
# print(active_users)
active_users = joiner(active_users)
# print(active_users)

fired_users = get_data('fired.xlsx', 2, 3722) # 3722
# print(fired_users)
# сравниваем списки
counter = 0
delete_list = []
for i in active_users:
    for j in fired_users:
        if i[1] == j[1]:
            delete_list.append(counter)
    counter += 1

# подготавливаем список элементов для удаления
final_delete_list = []
delete_list = set(delete_list)

for nums in delete_list:
    final_delete_list.append(nums)

final_delete_list.reverse()

# удаляем повторяющиеся элементы
for deleter in final_delete_list:
    active_users.pop(deleter)

print(final_delete_list)

# пишем результат в excel файл
book = op.Workbook()
sheet = book.active
row = 1
for items in active_users:
    column = 1
    data_cell_num = 0
    while column <= len(items):
        sheet.cell(row=row, column=column).value = items[data_cell_num]
        column += 1
        data_cell_num += 1
    row += 1
print(len(final_delete_list))
book.save('result_without_blckd-2.xlsx')
book.close()
