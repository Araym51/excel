from docx import Document
import openpyxl as op


def get_data(file_name, min_row, max_row):
    """
    функция принимает имя файла, номера строк которые нужно обработать,
    и возвращает список
    :param file_name: ередается имя файла (с расширением) для прочтения
    :param min_row: строка с которой начинается чтение
    :param max_row: строка на которой чтение файла заканчивается
    :return:
    """
    data_list = []
    sheets = op.load_workbook(filename=file_name).active
    for rows in sheets.iter_rows(min_row=min_row, max_row=max_row, min_col=0, max_col=9):
        x = []
        for cell in rows:
            x.append(cell.value)
        data_list.append(x)
    return data_list


# задаем имя файлов в качвычках и нужные нам строки min_row и max_row
source_list = get_data('source.xlsx', 2124, 2125)
document = Document()
n = 0

for pc_name, hdd_number, hdd_serial, pc_number in source_list:
    n += 1
    table = document.add_table(rows=4, cols=4)
    cell = table.cell(0, 0)
    cell.text = f'{n}.'
    cell = table.cell(0, 1)
    cell.text = f'АРМ с именем «{str(pc_name)}» в составе:'
    cell = table.cell(1, 1)
    cell.text = str('Системный блок')
    cell = table.cell(1, 2)
    cell.text = str('-')
    cell = table.cell(1, 3)
    if pc_number is None:
        cell.text = str(f'-')
    else:
        cell.text = str(f'{pc_number}')
    cell = table.cell(2, 1)
    if hdd_number is None:
        cell.text = str('-')
    else:
        cell.text = str(f'ЖМД № {hdd_number}')
    cell = table.cell(2, 2)
    cell.text = str('-')
    cell = table.cell(2, 3)
    if hdd_serial is None:
        cell.text = str('-')
    else:
        cell.text = str(hdd_serial)
    cell = table.cell(3, 1)
    cell.text = f'Периферийное оборудование: монитор, клавиатура, манипулятор мышь,' \
                f' принтер,акустические колонки, web-камера '
    a = table.cell(0, 1)
    b = table.cell(0, 3)
    a.merge(b)
    a = table.cell(3, 1)
    b = table.cell(3, 3)
    a.merge(b)
    a = table.cell(0, 0)
    b = table.cell(3, 0)
    table = a.merge(b)

document.save('ЦВСНП.docx')
