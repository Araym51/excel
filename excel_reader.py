import openpyxl as op


def get_data(file_name, min_row, max_row):
    """
    функция принимает имя файла, номера строк которые нужно обработать,
    и возвращает список
    :param file_name: ередается имя файла (с расширением) для прочтения
    :param min_row: строка с которой начинается чтение
    :param max_row: строка на которой чтение файла заканчивается
    :return:
    """
    data_list = []
    sheets = op.load_workbook(filename=file_name).active
    for rows in sheets.iter_rows(min_row=min_row, max_row=max_row, min_col=0, max_col=10):
        x = []
        for cell in rows:
            x.append(cell.value)
        data_list.append(x)
    return data_list


# задаем имя файлов в качвычках и нужные нам строки min_row и max_row
source_list = get_data('all.xlsx', 2415, 2498)
compare_list = get_data('LOVDT.xlsx', 2, 78)

# сравниваем списки
counter = 0
delete_list = []
for i in compare_list:
    for j in source_list:
        if i[5] == j[5]:
            delete_list.append(counter)
    counter += 1

# подготавливаем список элементов для удаления
final_delete_list = []
delete_list = set(delete_list)

for nums in delete_list:
    final_delete_list.append(nums)

final_delete_list.reverse()

# удаляем повторяющиеся элементы
for deleter in final_delete_list:
    compare_list.pop(deleter)

# пишем результат в excel файл
book = op.Workbook()
sheet = book.active
row = 1
for items in compare_list:
    column = 1
    while column < len(items):
        sheet.cell(row=row, column=column).value = items[column]
        column += 1
    row += 1

book.save('result.xlsx')
book.close()


if __name__ == "__main__":
    pass