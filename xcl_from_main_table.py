# pip install openpyxl
import openpyxl as op
from docx import Document

def get_data(file_name, min_row, max_row):
    """
    функция принимает имя файла, номера строк которые нужно обработать,
    и возвращает список
    :param file_name: ередается имя файла (с расширением) для прочтения
    :param min_row: строка с которой начинается чтение
    :param max_row: строка на которой чтение файла заканчивается
    :return:
    """
    data_list = []
    sheets = op.load_workbook(filename=file_name).active
    for rows in sheets.iter_rows(min_row=min_row, max_row=max_row, min_col=0, max_col=9):
        x = []
        for cell in rows:
            x.append(cell.value)
        data_list.append(x)
    return data_list


def table_writer(source, os, name):
    document = Document()
    file_name = name
    n = 0
    if len(source) == 0:
        pass
    else:
        for item in source:
                n += 1
                table = document.add_table(rows=4, cols=4)
                cell = table.cell(0, 0)
                cell.text = f'{n}.'
                cell = table.cell(0, 1)
                cell.text = f'АРМ с именем «{str(item[6])}» в составе:'
                cell = table.cell(1, 1)
                cell.text = str('Системный блок')
                cell = table.cell(1, 2)
                cell.text = str('-')
                cell = table.cell(1, 3)
                if item[4] is None:
                    cell.text = str(f'-')
                else:
                    cell.text = str(f'{item[4]}')
                cell = table.cell(2, 1)
                if item[2] is None:
                    cell.text = str('-')
                else:
                    cell.text = str(f'ЖМД № {item[2]}')
                cell = table.cell(2, 2)
                cell.text = str('-')
                cell = table.cell(2, 3)
                if item[3] is None:
                    cell.text = str('-')
                else:
                    cell.text = str(item[3])
                cell = table.cell(3, 1)
                cell.text = f'Периферийное оборудование: монитор, клавиатура, манипулятор мышь,' \
                            f' принтер,акустические колонки, web-камера '
                a = table.cell(0, 1)
                b = table.cell(0, 3)
                a.merge(b)
                a = table.cell(3, 1)
                b = table.cell(3, 3)
                a.merge(b)
                a = table.cell(0, 0)
                b = table.cell(3, 0)
                table = a.merge(b)

        if os == 'win':
            document.save(f'{file_name}.docx')
        elif os == 'lin':
            document.save(f'{file_name} - linux.docx')
        else:
            print('ОС не указана')
            exit(0)

source_list = get_data('all.xlsx', 123, 203)
with_linux = []
with_windows = []

for item in source_list:
    if item[8] != None:
        if 'lin' in str.lower(f'{item[8]}'):
            with_linux.append(item)
        else:
            with_windows.append(item)

table_writer(with_linux, 'lin', 'АЛЕКСАНДРОВСК')
table_writer(with_windows, 'win', 'АЛЕКСАНДРОВСК')
